import openpyxl as op
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from string import Template
from pathlib import Path
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import random


# Method to send emails


def send_mail(sname, recipient):

    # Enter your admin email and password here
    sender =  # " YourMail@mailer.com "
    password =  # " yourPasswordHere "
    source1 = # url or image 1
    source2 = # url or image 2
    source3 = # url or image 3
    source4 = # url or image 4

    n = random.randint(1, 4)
    if(n == 1):
        source = source1
    elif(n == 2):
        source = source2
    elif(n == 2):
        source = source3
    else:
        source = source4
    template = Template(Path("template.html").read_text())
    # image1 = MIMEImage(Path("image1.jpg").read_bytes())
    # image2 = MIMEImage(Path("image2.png").read_bytes())
    body = template.substitute(tag=source, name=sname)
    message = MIMEMultipart()
    message["from"] = "MIT MEERUT"
    message["to"] = recipient
    message["subject"] = "WISHES FROM MEERUT INSTITUTE OF TECHNOLOGY"
    message.attach(MIMEText(body, "html"))

    # message.add_header("Content_ID", "< img1 >")
    # message.add_header("Content_ID", "< img2 >")
    # message.attach(image1)
    # message.attach(image2)

    with smtplib.SMTP(host="smtp.gmail.com", port=587) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.login(sender, password)
        smtp.send_message(message)
        return True


# scope = ['https://spreadsheets.google.com/feeds',
#          'https://www.googleapis.com/auth/drive']
# creds = ServiceAccountCredentials.from_json_keyfile_name(
#     'client_secret.json', scope)
# client = gspread.authorize(creds)
# sheet = client.open("details").sheet1

wb = op.load_workbook("MITAlumni.xlsx")

sheet = wb["MITAlumni"]


i = 0
# Iterating over spreadsheet
while True:
    i += 1
    sdate = str(sheet.cell(i, 4).value)
    # print("sys date", str(date.today().strftime("%m-%d")))
    if(sdate[5:10] == str(date.today().strftime("%m/%d")) or sdate[0:5] == str(date.today().strftime("%d/%m"))):
        bool = send_mail(sheet.cell(i, 1).value, sheet.cell(i, 3).value)
        if(bool == True):
            print("mail send to {} at {}".format(
                str(sheet.cell(i, 3).value), str(sheet.cell(i, 6).value)))
    if(sheet.cell(i, 3).value == None):
        break
